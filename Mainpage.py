import streamlit as st
import io
import pandas as pd
from openpyxl import load_workbook
from odf.opendocument import load
from odf.text import P, Span, H, ListItem
import time
from datetime import datetime, timedelta
import zipfile

# 이 파일이 직접 실행될 때만 페이지 설정을 적용합니다.
if __name__ == "__main__":
    st.set_page_config(layout="wide")

st.title("급수공사 자동화 세션 💧")
st.subheader("내역서 ➡️ 집행건의 or 승인공문")

def number_to_korean(n):
    units = ['', '만', '억', '조', '경', '해']
    nums = [''] + list('일이삼사오육칠팔구')
    num_str = str(n)[::-1]
    parts = [num_str[i:i+4][::-1] for i in range(0, len(num_str), 4)]
    korean_str = ''
    
    for i, part in enumerate(parts):
        part_str = ''
        for j, digit in enumerate(part[::-1]):
            if digit != '0':
                part_str = nums[int(digit)] + (['십', '백', '천'][j-1] if j > 0 else '') + part_str
        if part_str:
            korean_str = part_str + units[i] + korean_str

    return korean_str or '영'

def format_with_commas(number):
    return "{:,}".format(number)

def extract_excel_values(sheet):
    amount1 = format_with_commas(sheet['N24'].value) if sheet['N24'].value else "-"
    amount2 = format_with_commas(sheet['N19'].value) if sheet['N19'].value else "-"
    amount3 = format_with_commas(sheet['N20'].value) if sheet['N20'].value else "-"
    amount4 = format_with_commas(sheet['N22'].value) if sheet['N22'].value else "-"
    amount5 = format_with_commas(sheet['N23'].value) if sheet['N23'].value else "-"
    
    address = sheet['D4'].value
    if address:
        address = address.replace("성남시 분당구", "").replace("성남시 중원구", "").replace("성남시 수정구", "").strip()
    
    consumer = sheet['M4'].value
    
    sentence = ""
    diameter = ""
    for col in range(4, 15):
        cell_value = sheet.cell(row=2, column=col).value
        if isinstance(cell_value, (int, float)):
            diameter = str(int(cell_value)) + "mm"  # mm 단위 추가
            usage = sheet.cell(row=1, column=col).value
            count = str(sheet.cell(row=3, column=col).value)
            
            usage_inside = usage.split('(')[-1].replace(")", "")
            usage_outside = usage.split('(')[0].strip() if '(' in usage else usage
            
            sentence = f"{usage_inside}용 계량기 {diameter} {count}전 {usage_outside}"
            break
    
    return amount1, amount2, amount3, amount4, amount5, address, consumer, sentence, diameter

def replace_text_in_node(node, search_text, replace_text):
    if node.nodeType == node.TEXT_NODE:
        if search_text in node.data:
            node.data = node.data.replace(search_text, str(replace_text))
    else:
        for child in node.childNodes:
            replace_text_in_node(child, search_text, replace_text)

def replace_text_in_elements(doc, replacements):
    for search_text, replace_text in replacements.items():
        elements = [P, Span, H, ListItem]
        for elem in elements:
            for node in doc.getElementsByType(elem):
                replace_text_in_node(node, search_text, replace_text)

def create_approval_document(amount4, address, consumer, sentence, 공사유형, amount2, amount3, input_date):
    if amount4 == "-":
        filepath = "C:/Users/Owner/Desktop/Gnudas GPT/서식/승인.odt"
    else:
        filepath = "C:/Users/Owner/Desktop/Gnudas GPT/서식/승인원인자.odt"

    try:
        with open(filepath, "rb") as file:
            doc = load(io.BytesIO(file.read()))
        
        future_date = input_date + timedelta(days=25)
        formatted_date = future_date.strftime("%Y. %m. %d.")

        replacements = {
            "입력주소": address,
            "입력수용가": consumer,
            "문구": sentence,
            "공사유형": 공사유형,
            "입력금액1": amount1,  # 주의: amount1 변수가 전역에서 선언되지 않았으므로 필요에 따라 수정하세요.
            "입력금액4": amount4,
            "입력금액5": amount5,
            "입력금액6": format_with_commas(int(amount2.replace(',', '')) + int(amount3.replace(',', '')))
                         if amount2 != "-" and amount3 != "-" else "-",
            "입력내용": sentence,
        }
        
        if amount4 != "-":
            replacements["입력일자"] = formatted_date

        replace_text_in_elements(doc, replacements)

        out_file = io.BytesIO()
        doc.save(out_file)
        out_file.seek(0)

        approval_file_name = "승인_공문.odt" if amount4 == "-" else "승인원인자_공문.odt"

        cause_excel_path = "C:/Users/Owner/Desktop/Gnudas GPT/서식/원인자엑셀.xlsx"
        wb_cause = load_workbook(cause_excel_path)
        ws_cause = wb_cause.active
        
        ws_cause['B8'] = address
        ws_cause['C8'] = consumer
        ws_cause['E8'] = sentence
        ws_cause['G8'] = amount1
        ws_cause['H8'] = format_with_commas(int(amount2.replace(',', '')) + int(amount3.replace(',', '')))
                         if amount2 != "-" and amount3 != "-" else "-"
        ws_cause['I8'] = amount4
        ws_cause['J8'] = amount5
        ws_cause['C27'] = diameter.replace("mm", "")

        excel_out_file = io.BytesIO()
        wb_cause.save(excel_out_file)
        excel_out_file.seek(0)

        return out_file, approval_file_name, excel_out_file
    except Exception as e:
        st.error(f"오류가 발생했습니다: {e}")

# 엑셀 파일 업로드 UI
uploaded_file = st.file_uploader("엑셀 파일을 업로드하세요", type=["xlsm", "xlsx"])

if uploaded_file is not None:
    progress_bar = st.progress(0)
    for i in range(100):
        time.sleep(0.02)
        progress_bar.progress(i + 1)

    wb = load_workbook(uploaded_file, data_only=True, keep_vba=True)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if "급수공사" in sheet and not ws.sheet_state == "hidden":
            st.write(f"활성화된 시트: {sheet}")

            amount1, amount2, amount3, amount4, amount5, address, consumer, sentence, diameter = extract_excel_values(ws)
            
            amount_korean = number_to_korean(int(amount1.replace(',', ''))) if amount1 != "-" else "-"
            
            st.write(f"주소: {address}")
            st.write(f"금액1: {amount1}, 금액2: {amount2}, 금액3: {amount3}, 금액4: {amount4}, 금액5: {amount5}")
            st.write(f"수용가: {consumer}")
            st.write(f"문구: {sentence}")

            공사유형 = "신설" if "신설" in sentence else "개조"
            st.write(f"공사유형: {공사유형}")

            if st.button("집행건의 생성"):
                original_filepath = "C:/Users/Owner/Desktop/Gnudas GPT/서식/급수공사 집행 건의 서식.odt"
                try:
                    with open(original_filepath, "rb") as file:
                        odt_file = file.read()
                    doc = load(io.BytesIO(odt_file))

                    replacements = {
                        "입력주소": address,
                        "입력금액1": amount1,
                        "입력금액2": amount2,
                        "입력금액3": amount3,
                        "입력금액4": amount4,
                        "입력금액5": amount5,
                        "입력금액한글": amount_korean,
                        "입력수용가": consumer,
                        "입력내용": sentence,
                        "공사유형": 공사유형,
                        "입력금액6": format_with_commas(int(amount2.replace(',', '')) + int(amount3.replace(',', '')))
                                      if amount2 != "-" and amount3 != "-" else "-"
                    }
                    
                    replace_text_in_elements(doc, replacements)

                    out_file = io.BytesIO()
                    doc.save(out_file)
                    out_file.seek(0)

                    odt_file_name = f"급수공사 집행건의({address}).odt"

                    st.download_button(
                        label="⬇️ 집행건의 다운로드",
                        data=out_file,
                        file_name=odt_file_name,
                        mime="application/vnd.oasis.opendocument.text"
                    )
                    st.success("ODT 파일이 성공적으로 생성되었습니다.")
                except Exception as e:
                    st.error(f"오류가 발생했습니다: {e}")

            if st.button("승인 공문 생성"):
                input_date = datetime.today()
                out_file, approval_file_name, excel_out_file = create_approval_document(amount4, address, consumer, sentence, 공사유형, amount2, amount3, input_date)
                
                if out_file:
                    st.download_button(
                        label="⬇️ 승인 공문 다운로드",
                        data=out_file,
                        file_name=f"급수공사 승인({address}).odt",
                        mime="application/vnd.oasis.opendocument.text"
                    )

                    if excel_out_file:
                        st.download_button(
                            label="⬇️ 원인자 엑셀 다운로드",
                            data=excel_out_file,
                            file_name="원인자엑셀.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

            break
