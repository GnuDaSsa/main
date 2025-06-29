import streamlit as st
import io
import pandas as pd
from openpyxl import load_workbook
from odf.opendocument import load
from odf.text import P, Span, H, ListItem
import time
from datetime import datetime, timedelta
import zipfile
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def number_to_korean(n):
    units = ['', '만', '억', '조', '경', '해']
    nums = [''] + list('일이삼사오육칠팔구')
    num_str = str(n)[::-1]
    parts = [num_str[i:i+4][::-1] for i in range(0, len(num_str), 4)]
    korean_str = ''
    for i, part in enumerate(parts):
        part_str = ''
        for j, digit in enumerate(part[::-1]):
            if digit != '0':
                part_str = nums[int(digit)] + (['십', '백', '천'][j-1] if j > 0 else "") + part_str
        if part_str:
            korean_str = part_str + units[i] + korean_str
    return korean_str or '영'

def format_with_commas(number):
    return "{:,}".format(number)

def extract_excel_values(sheet):
    amount1 = format_with_commas(sheet['N24'].value) if sheet['N24'].value else "-"
    amount2 = format_with_commas(sheet['N19'].value) if sheet['N19'].value else "-"
    amount3 = format_with_commas(sheet['N20'].value) if sheet['N20'].value else "-"
    amount4 = format_with_commas(sheet['N22'].value) if sheet['N22'].value else "-"
    amount5 = format_with_commas(sheet['N23'].value) if sheet['N23'].value else "-"
    
    address = sheet['D4'].value
    if address:
        address = address.replace("성남시 분당구", "").replace("성남시 중원구", "").replace("성남시 수정구", "").strip()
    
    consumer = sheet['M4'].value
    sentence = ""
    diameter = ""
    for col in range(4, 15):
        cell_value = sheet.cell(row=2, column=col).value
        if isinstance(cell_value, (int, float)):
            diameter = str(int(cell_value)) + "mm"
            usage = sheet.cell(row=1, column=col).value
            count = str(sheet.cell(row=3, column=col).value)
            usage_inside = usage.split('(')[-1].replace(")", "")
            usage_outside = usage.split('(')[0].strip() if '(' in usage else usage
            sentence = f"{usage_inside}용 계량기 {diameter} {count}전 {usage_outside}"
            break
    return amount1, amount2, amount3, amount4, amount5, address, consumer, sentence, diameter

def replace_text_in_node(node, search_text, replace_text):
    if node.nodeType == node.TEXT_NODE:
        if search_text in node.data:
            node.data = node.data.replace(search_text, str(replace_text))
    else:
        for child in node.childNodes:
            replace_text_in_node(child, search_text, replace_text)

def replace_text_in_elements(doc, replacements):
    for search_text, replace_text in replacements.items():
        elements = [P, Span, H, ListItem]
        for elem in elements:
            for node in doc.getElementsByType(elem):
                replace_text_in_node(node, search_text, replace_text)

def create_approval_document(amount4, address, consumer, sentence, 공사유형, amount2, amount3, input_date, amount1, amount5, diameter):
    filepath = os.path.join(BASE_DIR, "서식", "page1", "승인.odt") if amount4 == "-" else os.path.join(BASE_DIR, "서식", "page1", "승인원인자.odt")
    try:
        with open(filepath, "rb") as file:
            doc = load(io.BytesIO(file.read()))
        
        future_date = input_date + timedelta(days=25)
        formatted_date = future_date.strftime("%Y. %m. %d.")
        
        replacements = {
            "입력주소": address,
            "입력수용가": consumer,
            "문구": sentence,
            "공사유형": 공사유형,
            "입력금액1": amount1,
            "입력금액4": amount4,
            "입력금액5": amount5,
            "입력금액6": format_with_commas(int(amount2.replace(',', '')) + int(amount3.replace(',', ''))) if amount2 != "-" and amount3 != "-" else "-",
            "입력내용": sentence,
        }
        if amount4 != "-":
            replacements["입력일자"] = formatted_date
        replace_text_in_elements(doc, replacements)
        
        out_file = io.BytesIO()
        doc.save(out_file)
        out_file.seek(0)
        
        approval_file_name = "승인_공문.odt" if amount4 == "-" else "승인원인자_공문.odt"
        
        cause_excel_path = os.path.join(BASE_DIR, "서식", "page1", "원인자엑셀.xlsx")
        wb_cause = load_workbook(cause_excel_path)
        ws_cause = wb_cause.active
        
        ws_cause['B8'] = address
        ws_cause['C8'] = consumer
        ws_cause['E8'] = sentence
        ws_cause['G8'] = amount1
        ws_cause['H8'] = format_with_commas(int(amount2.replace(',', '')) + int(amount3.replace(',', ''))) if amount2 != "-" and amount3 != "-" else "-"
        ws_cause['I8'] = amount4
        ws_cause['J8'] = amount5
        ws_cause['C27'] = diameter.replace("mm", "")
        
        excel_out_file = io.BytesIO()
        wb_cause.save(excel_out_file)
        excel_out_file.seek(0)
        
        return out_file, approval_file_name, excel_out_file
    except Exception as e:
        st.error(f"오류가 발생했습니다: {e}")

def run():
    # 페이지 설정 제거: app.py에서 이미 설정됨
    st.title("급수공사 자동화 세션 💧")
    st.subheader("내역서 ➡️ 공문 생성 (집행건의 & 승인 공문 통합)")
    
    uploaded_file = st.file_uploader("엑셀 파일을 업로드하세요", type=["xlsm", "xlsx"])
    
    if uploaded_file is not None:
        progress_bar = st.progress(0)
        for i in range(100):
            time.sleep(0.02)
            progress_bar.progress(i + 1)
        wb = load_workbook(uploaded_file, data_only=True, keep_vba=True)
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            if "급수공사" in sheet and not ws.sheet_state == "hidden":
                st.write(f"활성화된 시트: {sheet}")
                amount1, amount2, amount3, amount4, amount5, address, consumer, sentence, diameter = extract_excel_values(ws)
                amount_korean = number_to_korean(int(amount1.replace(',', ''))) if amount1 != "-" else "-"
                
                st.write(f"주소: {address}")
                st.write(f"금액1: {amount1}, 금액2: {amount2}, 금액3: {amount3}, 금액4: {amount4}, 금액5: {amount5}")
                st.write(f"수용가: {consumer}")
                st.write(f"문구: {sentence}")
                
                공사유형 = "신설" if "신설" in sentence else "개조"
                st.write(f"공사유형: {공사유형}")
                
                if st.button("공문 생성"):
                    try:
                        original_filepath = os.path.join(BASE_DIR, "서식", "page1", "급수공사 집행 건의 서식.odt")
                        with open(original_filepath, "rb") as file:
                            odt_file = file.read()
                        doc = load(io.BytesIO(odt_file))
                        replacements = {
                            "입력주소": address,
                            "입력금액1": amount1,
                            "입력금액2": amount2,
                            "입력금액3": amount3,
                            "입력금액4": amount4,
                            "입력금액5": amount5,
                            "입력금액한글": amount_korean,
                            "입력수용가": consumer,
                            "입력내용": sentence,
                            "공사유형": 공사유형,
                            "입력금액6": format_with_commas(int(amount2.replace(',', '')) + int(amount3.replace(',', ''))) if amount2 != "-" and amount3 != "-" else "-"
                        }
                        replace_text_in_elements(doc, replacements)
                        exec_odt_buffer = io.BytesIO()
                        doc.save(exec_odt_buffer)
                        exec_odt_buffer.seek(0)
                        exec_file_name = f"급수공사 집행건의({address}).odt"
                    except Exception as e:
                        st.error(f"집행건의 생성 오류: {e}")
                        exec_odt_buffer = None
                    
                    try:
                        input_date = datetime.today()
                        approval_result = create_approval_document(amount4, address, consumer, sentence, 공사유형, amount2, amount3, input_date, amount1, amount5, diameter)
                        if approval_result:
                            approval_odt_buffer, approval_file_name, excel_buffer = approval_result
                        else:
                            approval_odt_buffer, approval_file_name, excel_buffer = None, None, None
                    except Exception as e:
                        st.error(f"승인 공문 생성 오류: {e}")
                        approval_odt_buffer, approval_file_name, excel_buffer = None, None, None
                    
                    if exec_odt_buffer and approval_odt_buffer:
                        zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
                            zipf.writestr(exec_file_name, exec_odt_buffer.getvalue())
                            zipf.writestr(approval_file_name, approval_odt_buffer.getvalue())
                            if amount4 != "-" and excel_buffer:
                                zipf.writestr("원인자엑셀.xlsx", excel_buffer.getvalue())
                        zip_buffer.seek(0)
                        st.download_button(
                            label="⬇️ 공문 전체 다운로드",
                            data=zip_buffer,
                            file_name="공문.zip",
                            mime="application/zip"
                        )
                        st.success("모든 공문이 성공적으로 생성되었습니다.")
                    else:
                        st.error("파일 생성에 실패하였습니다.")
                break

if __name__ == "__main__":
    run()