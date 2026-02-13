import streamlit as st
import openpyxl
import pandas as pd
from datetime import datetime
import io
import zipfile
import tempfile
import os
import platform

# MongoDB 및 ODT 관련 라이브러리 import
from odf.opendocument import load
from odf.element import Element, Text
from mongo_env import get_mongo_uri, get_collection

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def run():
    st.title("도급위탁용역 점검표 생성기 (ODT)")
    
    if not get_mongo_uri():
        st.error("MongoDB 연결 정보(MONGODB_URI)가 설정되지 않았습니다.")
        st.info("로컬: `.env` 설정 / 배포: Streamlit Secrets 설정을 추가하세요.")
        return
    
    if platform.system() == "Windows":
        base_path = os.path.join(os.path.expanduser('~'), "Desktop", "사진우", "AI", "도급위탁용역자동화")
    else:
        base_path = os.path.join(os.path.expanduser('~'), "projects", "AIworks", "도급위탁용역자동화")
    template_paths_by_sheet = {
        "점검표1": os.path.join(BASE_DIR, "서식", "page4", "점검표(60일 이상).odt"),
        "점검표2": os.path.join(BASE_DIR, "서식", "page4", "점검표(60일 이하).odt"),
        "점검표3": os.path.join(BASE_DIR, "서식", "page4", "점검표3.odt")
    }

    full_compliance_groups = [
        (17, 18, 19, ["[Q]", "[R]", "[S]"]), (20, 21, 22, ["[T]", "[U]", "[V]"]),
        (23, 24, 25, ["[W]", "[X]", "[Y]"]), (26, 27, 28, ["[Z]", "[AA]", "[AB]"]),
        (29, 30, 31, ["[AC]", "[AD]", "[AE]"]), (32, 33, 34, ["[AF]", "[AG]", "[AH]"]),
        (35, 36, 37, ["[AI]", "[AJ]", "[AK]"]), (38, 39, 40, ["[AL]", "[AM]", "[AN]"]),
        (41, 42, 43, ["[AO]", "[AP]", "[AQ]"]), (44, 45, 46, ["[AR]", "[AS]", "[AT]"])
    ]
    compliance_groups_by_sheet = {
        "점검표1": full_compliance_groups, "점검표2": full_compliance_groups[:7], "점검표3": full_compliance_groups
    }
    st.markdown("""
    <style>
    div.stTextInput > div > input { max-width: 200px; }
    [data-baseweb="input"] > input::placeholder { color: #999 !important; }
    </style>
    """, unsafe_allow_html=True)
    if "extracted_page4" not in st.session_state: st.session_state.extracted_page4 = []
    if "apply_all_checkbox_page4" not in st.session_state: st.session_state.apply_all_checkbox_page4 = False

    left_col, right_col = st.columns([1, 1])

    with left_col:
        c1, c2 = st.columns([1, 1]);
        with c1: team_leader = st.text_input("팀장님 성함", max_chars=10, key="team_leader_page4")
        with c2: manager = st.text_input("과장님 성함", max_chars=10, key="manager_page4")
        uploaded_file = st.file_uploader("엑셀 파일을 업로드하세요", type=["xlsx"], key="uploaded_file_page4")
        generate_btn = st.button("생성", key="generate_btn_page4")

        if generate_btn and (not team_leader or not manager): st.warning("팀장님과 과장님 성함을 모두 입력하세요.")
        elif generate_btn and not uploaded_file: st.warning("엑셀 파일을 업로드하세요.")

        if generate_btn and team_leader and manager and uploaded_file:
            def extract_data_from_sheets():
                wb = openpyxl.load_workbook(uploaded_file)
                target_sheet_names_prefixes = ["점검표1", "점검표2", "점검표3"]
                all_extracted_data_flat, extracted_data_by_sheet = [], {}
                for sheet_prefix in target_sheet_names_prefixes:
                    actual_sheet_name = next((s for s in wb.sheetnames if sheet_prefix in s), None)
                    if not actual_sheet_name:
                        st.warning(f"'{sheet_prefix}' 시트를 찾을 수 없습니다."); continue
                    sh = wb[actual_sheet_name]
                    current_sheet_extracted_data = []
                    sheet_compliance_groups = compliance_groups_by_sheet.get(sheet_prefix, full_compliance_groups)
                    for row in sh.iter_rows(min_row=2, max_row=sh.max_row, min_col=2, max_col=2):
                        cell = row[0]
                        if cell.value and cell.font and cell.font.color and cell.font.color.rgb == "FFFF0000":
                            def fmt_date(d): return d.strftime("%Y.%m.%d.") if isinstance(d, datetime) else d
                            def fmt_cost(c): return f"{int(c):,}" if isinstance(c, (int, float)) else c
                            mgr_raw = sh.cell(row=cell.row, column=6).value or ""
                            row_data = {
                                "사업명": sh.cell(row=cell.row, column=2).value, "부서명": sh.cell(row=cell.row, column=5).value,
                                "담당자": mgr_raw.split('(')[0].strip(), "착공일": fmt_date(sh.cell(row=cell.row, column=8).value),
                                "준공일": fmt_date(sh.cell(row=cell.row, column=9).value), "사업비": fmt_cost(sh.cell(row=cell.row, column=11).value),
                                "팀장님": team_leader, "과장님": manager, "source_sheet": sheet_prefix,
                                "created_at": datetime.now() # 데이터 생성 시각 추가
                            }
                            m = str(sh.cell(row=cell.row, column=13).value).strip().upper()
                            row_data["M1"], row_data["M2"] = ("○", "") if m == "O" else ("", "○") if m == "X" else ("", "")
                            for i, (c1, c2, c3, tags) in enumerate(sheet_compliance_groups, 1):
                                v1, v2, v3 = sh.cell(row=cell.row, column=c1).value, sh.cell(row=cell.row, column=c2).value, sh.cell(row=cell.row, column=c3).value
                                row_data[f"의무{i}"] = tags[0] if v1 else tags[1] if v2 else tags[2] if v3 else None
                            current_sheet_extracted_data.append(row_data)
                    if current_sheet_extracted_data:
                        all_extracted_data_flat.extend(current_sheet_extracted_data)
                        extracted_data_by_sheet[sheet_prefix] = current_sheet_extracted_data
                return all_extracted_data_flat, extracted_data_by_sheet
            
            st.session_state.extracted_page4, st.session_state.extracted_data_by_sheet_page4 = extract_data_from_sheets()

            if not st.session_state.extracted_page4:
                st.warning("조건을 만족하는 데이터가 없습니다.")
            else:
                for sheet_name, data_list in st.session_state.extracted_data_by_sheet_page4.items():
                    st.subheader(f"■ {sheet_name} 추출값"); st.dataframe(pd.DataFrame(data_list), use_container_width=True)
                for i in range(len(st.session_state.extracted_page4)):
                    if f"date1_page4_{i}" not in st.session_state: st.session_state[f"date1_page4_{i}"] = ""
                    if f"date2_page4_{i}" not in st.session_state: st.session_state[f"date2_page4_{i}"] = ""

                # MongoDB 저장 로직
                try:
                    all_data_to_save = st.session_state.extracted_page4
                    st.info("MongoDB에 연결 중...")
                    collection = get_collection("automation_db", "inspection_records")
                    if collection is None:
                        st.error("MongoDB 연결에 실패했습니다. URI 및 네트워크 설정을 확인하세요.")
                    else:
                        st.info(f"{len(all_data_to_save)}개의 데이터를 저장합니다...")
                        result = collection.insert_many(all_data_to_save)
                        st.success(f"{len(result.inserted_ids)}개의 데이터가 MongoDB에 성공적으로 저장되었습니다.")
                except Exception as e:
                    st.error(f"데이터를 저장하는 중 오류가 발생했습니다: {e}")
    
    with left_col:
        if st.session_state.get("extracted_page4"):
            c0, c1, c2, c3 = st.columns([2, 1, 1, 0.5]); first_item = st.session_state.extracted_page4[0]
            with c0: st.write(f"■ {first_item['사업명']}")
            with c1: st.text_input("점검일자", key="date1_page4_0", placeholder="YYYYMMDD")
            with c2: st.text_input("준공검사일", key="date2_page4_0", placeholder="YYYYMMDD")
            with c3: st.markdown("<div style='margin-top: 30px;'></div>", unsafe_allow_html=True); st.checkbox("일괄 적용", key="apply_all_checkbox_page4")
            if st.session_state.apply_all_checkbox_page4:
                d1, d2 = st.session_state.date1_page4_0, st.session_state.date2_page4_0
                for i in range(1, len(st.session_state.extracted_page4)): st.session_state[f'date1_page4_{i}'], st.session_state[f'date2_page4_{i}'] = d1, d2
            for idx in range(1, len(st.session_state.extracted_page4)):
                item = st.session_state.extracted_page4[idx]; c0, c1, c2, _ = st.columns([2, 1, 1, 0.5])
                with c0: st.write(f"■ {item['사업명']}")
                with c1: st.text_input("점검일자", key=f"date1_page4_{idx}", label_visibility="collapsed", value=st.session_state.get(f'date1_page4_{idx}', ''))
                with c2: st.text_input("준공검사일", key=f"date2_page4_{idx}", label_visibility="collapsed", value=st.session_state.get(f'date2_page4_{idx}', ''))

            if st.button("최종 생성", key="final_generate_btn_page4"):
                def find_and_replace_text(element, find_str, replace_str):
                    from odf.element import Element, Text
                    for child in list(element.childNodes):
                        if isinstance(child, Text): child.data = child.data.replace(find_str, replace_str)
                        elif isinstance(child, Element): find_and_replace_text(child, find_str, replace_str)
                buffer = io.BytesIO()
                with zipfile.ZipFile(buffer, "w") as zipf:
                    for idx, data in enumerate(st.session_state.extracted_page4):
                        template_path = template_paths_by_sheet.get(data["source_sheet"])
                        if not template_path or not os.path.exists(template_path):
                            st.error(f"템플릿을 찾을 수 없습니다: {template_path}. 건너뜁니다."); continue
                        doc = load(template_path)
                        def fmt_odt_date(s): return f"{s[:4]}. {int(s[4:6])}. {int(s[6:8])}." if len(s) == 8 and s.isdigit() else s
                        replacements = {f"[{k}]": str(v) for k, v in data.items() if k not in ["source_sheet", "created_at"]}
                        replacements["[점검일자]"] = fmt_odt_date(st.session_state.get(f'date1_page4_{idx}', ""))
                        replacements["[준공검사일]"] = fmt_odt_date(st.session_state.get(f'date2_page4_{idx}', ""))
                        current_groups = compliance_groups_by_sheet.get(data["source_sheet"], full_compliance_groups)
                        for i, (_, _, _, tags) in enumerate(current_groups, 1):
                            tag_to_mark = data.get(f"의무{i}")
                            for tag in tags: replacements[tag] = "○" if tag == tag_to_mark else ""
                        for find, replace in replacements.items():
                            if find: find_and_replace_text(doc.text, find, str(replace))
                        safe_name = "".join(c for c in data['사업명'] if c.isalnum() or c in ' _-').strip() or f"점검표_{idx+1}"
                        filename = f"[{data['source_sheet']}]{safe_name}.odt"
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".odt") as tmp:
                            doc.save(tmp.name)
                            zipf.write(tmp.name, arcname=filename)
                        os.unlink(tmp.name)
                buffer.seek(0)
                st.download_button("ZIP 다운로드", data=buffer, file_name="점검표_결과.zip", mime="application/zip")

    with right_col:
        image_path = os.path.join(BASE_DIR, "서식", "page4", "guide.png")
        if os.path.exists(image_path): st.image(image_path, use_container_width=True)
        else: st.error(f"설명 이미지를 찾을 수 없습니다: {image_path}")

if __name__ == "__main__":
    st.set_page_config(layout="wide", page_title="도급위탁용역 점검표 생성기")
    run()
